from dataclasses import dataclass, field, asdict
from datetime import datetime
import json
import openpyxl
from typing import Optional

@dataclass
class TrafficLog:
    filename: str
    resolution: Optional[str] = "NA"
    duration: Optional[float] = 0.0
    fps: Optional[float] = 0.0
    total_frames: Optional[int] = 0
    date_time_of_recording: Optional[str] = "NA"
    run_timestamp: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    source: str = "MISSING"
    location: str = "MISSING"
    camera_placement: str = "MISSING"
    camera_view_dirA: str = "MISSING"
    camera_fov: str = "MISSING"
    camera_elevation: str = "MISSING"
    weather: str = "MISSING"
    other_conditions: str = "MISSING"

    model_used: str = "NA"
    tracking_algorithm: str = "NA"
    accuracy: Optional[float] = None
    inference_time_per_frame_ms: Optional[float] = None
    false_positive_timestamps: str = ""
    average_fps: Optional[float] = None
    frame_drop_count: Optional[int] = 0
    notes: str = ""

    congestion_density_total: Optional[float] = None
    vpm_total: Optional[float] = None
    total_vehicle_count: Optional[int] = None
    car_count: Optional[int] = None
    truck_count: Optional[int] = None
    bus_count: Optional[int] = None
    motorcycle_count: Optional[int] = None

    congestion_density_dirA: Optional[float] = None
    vpm_dirA: Optional[float] = None
    dirA_vehicle_count: Optional[int] = None
    dirA_car_count: Optional[int] = None
    dirA_truck_count: Optional[int] = None
    dirA_bus_count: Optional[int] = None
    dirA_motorcycle_count: Optional[int] = None

    congestion_density_dirB: Optional[float] = None
    vpm_dirB: Optional[float] = None
    dirB_vehicle_count: Optional[int] = None
    dirB_car_count: Optional[int] = None
    dirB_truck_count: Optional[int] = None
    dirB_bus_count: Optional[int] = None
    dirB_motorcycle_count: Optional[int] = None

    overall_avg_speed: Optional[float] = None
    avg_car_speed: Optional[float] = None
    avg_truck_speed: Optional[float] = None
    avg_bus_speed: Optional[float] = None
    avg_motorcycle_speed: Optional[float] = None

    avg_speed_dirA: Optional[float] = None
    avg_car_speed_dirA: Optional[float] = None
    avg_truck_speed_dirA: Optional[float] = None
    avg_bus_speed_dirA: Optional[float] = None
    avg_motorcycle_speed_dirA: Optional[float] = None

    avg_speed_dirB: Optional[float] = None
    avg_car_speed_dirB: Optional[float] = None
    avg_truck_speed_dirB: Optional[float] = None
    avg_bus_speed_dirB: Optional[float] = None
    avg_motorcycle_speed_dirB: Optional[float] = None

    def update_from_metadata(self, metadata_dict):
        video_meta = metadata_dict.get(self.filename, {})
        self.source = video_meta.get("Source", self.source)
        self.location = video_meta.get("Location", self.location)
        self.camera_placement = video_meta.get("Camera Placement", self.camera_placement)
        self.camera_view_dirA = video_meta.get("Camera View (DirA)", self.camera_view_dirA)
        self.camera_fov = video_meta.get("Camera FOV", self.camera_fov)
        self.camera_elevation = video_meta.get("Camera Elevation", self.camera_elevation)
        self.weather = video_meta.get("Weather", self.weather)
        self.other_conditions = video_meta.get("Other Conditions", self.other_conditions)
        self.date_time_of_recording = video_meta.get("Date-Time of Recording", self.date_time_of_recording)

    def to_excel_row(self):
        def safe(val):
            return "NA" if val is None else val

        return [
            safe(self.filename), safe(self.resolution), safe(self.duration), safe(self.fps), safe(self.total_frames),
            safe(self.date_time_of_recording), safe(self.source), safe(self.location), safe(self.camera_placement),
            safe(self.camera_view_dirA), safe(self.camera_fov), safe(self.camera_elevation), safe(self.weather),
            safe(self.other_conditions), safe(self.run_timestamp), safe(self.model_used), safe(self.tracking_algorithm),
            safe(self.accuracy), safe(self.inference_time_per_frame_ms), safe(self.false_positive_timestamps),
            safe(self.average_fps), safe(self.frame_drop_count), safe(self.notes),
            safe(self.congestion_density_total), safe(self.vpm_total),
            safe(self.total_vehicle_count), safe(self.car_count), safe(self.truck_count), safe(self.bus_count), safe(self.motorcycle_count),
            safe(self.congestion_density_dirA), safe(self.vpm_dirA), safe(self.dirA_vehicle_count), safe(self.dirA_car_count),
            safe(self.dirA_truck_count), safe(self.dirA_bus_count), safe(self.dirA_motorcycle_count),
            safe(self.congestion_density_dirB), safe(self.vpm_dirB), safe(self.dirB_vehicle_count), safe(self.dirB_car_count),
            safe(self.dirB_truck_count), safe(self.dirB_bus_count), safe(self.dirB_motorcycle_count),
            safe(self.overall_avg_speed), safe(self.avg_car_speed), safe(self.avg_truck_speed), safe(self.avg_bus_speed), safe(self.avg_motorcycle_speed),
            safe(self.avg_speed_dirA), safe(self.avg_car_speed_dirA), safe(self.avg_truck_speed_dirA), safe(self.avg_bus_speed_dirA), safe(self.avg_motorcycle_speed_dirA),
            safe(self.avg_speed_dirB), safe(self.avg_car_speed_dirB), safe(self.avg_truck_speed_dirB), safe(self.avg_bus_speed_dirB), safe(self.avg_motorcycle_speed_dirB)
        ]

def load_video_metadata(json_path):
    with open(json_path, "r") as file:
        return json.load(file)

from openpyxl.styles import Alignment

def append_log_to_excel(log: TrafficLog, excel_path: str):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    row = log.to_excel_row()
    ws.append(row)
    row_idx = ws.max_row
    for cell in ws[row_idx]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(excel_path)
